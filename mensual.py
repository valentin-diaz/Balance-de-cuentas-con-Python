from openpyxl import Workbook, load_workbook
from statistics import mode
from time import sleep


# def formato_moneda(num):
#     # return "${:,.0f}".format(num)
#     return num

# Mensajes al usuario
print("Bienvenido. Recuerda abrir y guardar el archivo excel antes y después de correr este script.")

# Cargar el archivo y las hojas
wb_original = load_workbook("Cuentas.xlsx", data_only=False)
wb_trabajo = load_workbook("Cuentas.xlsx", data_only=True)
mov_s = wb_trabajo[wb_trabajo.sheetnames[0]]
mes_s = wb_trabajo[wb_trabajo.sheetnames[1]]

# Trabajar en C desde 5 hasta la última fila
# Decidir en qué mes y año se va a trabajar
# Recorrer C hasta encontrar la primera entrada de ese mes y año,
# y guardar el tipo y saldo final (lista 1) y la categoría (lista 2)

# Recorrer la lista 1:
# Si el tipo es entrada, subir el contador de entradas (o viceversa)
# Si es el primer o el último elemento, guardarlo
# Encontrar la moda de la lista 2

# Generar una nueva fila en mes_s con:
# Contador de entradas
# Contador de salidas
# Suma de ambos contadores
# Primer elemento de lista 1
# Último elemento de lista 1
# Moda de lista 2

# Cambiar el formato de E y F en la última fila

# Decidir mes y año
month = input("Indica el número de mes para realizar el balance: ")
year = int(input("Indica el año para realizar el balance: "))

# Lista con cada tipo de movimiento
tipos_mes = []
# Lista con saldos finales luego de cada movimiento
saldos_mes = []
# Lista con categoría de cada movimiento
cats_mes = []

# Loopear a través de la columna C y guardar los datos importantes
for row in range(5, mov_s.max_row + 1): 
    valor_celda = mov_s["C" + str(row)].value
    print(valor_celda)
    if f"-{month}-" in str(valor_celda) and f"{year}-" in str(valor_celda):
        print("ENTRA")
        tipo = mov_s["D" + str(row)].value
        saldo_final = mov_s["I" + str(row)].value
        categoria = mov_s["H" + str(row)].value
        tipos_mes.append(tipo)
        saldos_mes.append(saldo_final) 
        cats_mes.append(categoria)

# Contadores de entrada y salida
num_entradas = tipos_mes.count("Entrada")
num_salidas = tipos_mes.count("Salida")

# Categoría más común
cat_moda = mode(cats_mes)

# Añadir la fila a la página de excel
nueva_fila = [
    f"{month}-{year}",
    num_entradas, 
    num_salidas, 
    num_entradas + num_salidas, 
    saldos_mes[0], 
    saldos_mes[len(saldos_mes) - 1],
    saldos_mes[len(saldos_mes) - 1] - saldos_mes[0],
    cat_moda
    ]

print(nueva_fila)
mes_s_cambiar = wb_original[wb_original.sheetnames[1]]
mes_s_cambiar.append(nueva_fila)

# Cambiar el formato a moneda
# mes_s_cambiar["E" + str(mes_s_cambiar.max_row)].number_format = '"$"#,##0.00_);("$"#,##0.00)'

wb_original.save("Cuentas.xlsx")

print("El archivo fue actualizado con éxito. Ahora recuerda abrirlo y guardarlo manualmente.")
sleep(5)
